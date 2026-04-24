import re
import yaml
from pathlib import Path
from datetime import datetime, timezone
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[2]
SQL_ROOT = ROOT
DBT_ROOT = ROOT / "dbt"
OUTPUT_XLSX = ROOT / "FIP_Data_Dictionary.xlsx"
OUTPUT_MD = ROOT / "doc" / "DATA_DICTIONARY.md"

TABLE_RE = re.compile(r"CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?([a-zA-Z_][\w]*)\.([a-zA-Z_][\w]*)\s*\(", re.IGNORECASE)
VIEW_RE = re.compile(r"CREATE\s+(?:OR\s+REPLACE\s+)?VIEW\s+([a-zA-Z_][\w]*)\.([a-zA-Z_][\w]*)\s+AS", re.IGNORECASE)
COMMENT_ON_TABLE_RE = re.compile(r"COMMENT\s+ON\s+TABLE\s+([a-zA-Z_][\w]*)\.([a-zA-Z_][\w]*)\s+IS\s+\'([^\']*)\'", re.IGNORECASE)
COMMENT_ON_COLUMN_RE = re.compile(r"COMMENT\s+ON\s+COLUMN\s+([a-zA-Z_][\w]*)\.([a-zA-Z_][\w]*)\.([a-zA-Z_][\w]*)\s+IS\s+\'([^\']*)\'", re.IGNORECASE)


def _find_matching_paren(text: str, open_idx: int) -> int:
    depth = 0
    for i in range(open_idx, len(text)):
        ch = text[i]
        if ch == '(': depth += 1
        elif ch == ')':
            depth -= 1
            if depth == 0: return i
    return -1


def _split_top_level_columns(body: str):
    cols = []
    depth = 0
    start = 0
    for i, ch in enumerate(body):
        if ch == '(': depth += 1
        elif ch == ')': depth -= 1
        elif ch == ',' and depth == 0:
            cols.append(body[start:i].strip())
            start = i + 1
    tail = body[start:].strip()
    if tail: cols.append(tail)
    return cols


def _is_constraint_line(line: str) -> bool:
    s = line.strip().upper()
    return (
        s.startswith("CONSTRAINT ") or s.startswith("PRIMARY KEY") or
        s.startswith("FOREIGN KEY") or s.startswith("UNIQUE ") or
        s.startswith("CHECK ")
    )


def parse_sql_sources():
    objects = {}
    columns = []
    column_comments = {}

    for sql_file in sorted(SQL_ROOT.glob("fip_schema_*.sql")):
        text = sql_file.read_text(encoding="utf-8", errors="ignore")
        
        # Extract Tables
        for m in TABLE_RE.finditer(text):
            schema, name = m.group(1), m.group(2)
            open_idx = m.end() - 1
            close_idx = _find_matching_paren(text, open_idx)
            if close_idx == -1: continue
            
            obj_key = f"{schema}.{name}"
            objects[obj_key] = {
                "schema": schema,
                "name": name,
                "type": "TABLE",
                "description": "",
                "source": str(sql_file.name)
            }
            
            body = text[open_idx + 1:close_idx]
            ordinal = 0
            for entry in _split_top_level_columns(body):
                if not entry or _is_constraint_line(entry): continue
                
                # Extract column-level comment from DDL line
                col_comment = ""
                if "--" in entry:
                    entry_parts = entry.split("--", 1)
                    col_comment = entry_parts[1].strip()
                    line = entry_parts[0].strip()
                else:
                    line = entry.strip()

                if not line: continue
                parts = re.split(r"\s+", line, maxsplit=2)
                if len(parts) < 2: continue
                
                col_name = parts[0].strip('"[]')
                data_type = parts[1].strip(',')
                nullable = "NO" if "NOT NULL" in line.upper() else "YES"
                ordinal += 1
                columns.append({
                    "schema": schema,
                    "table": name,
                    "column": col_name,
                    "type": data_type,
                    "nullable": nullable,
                    "description": col_comment, # Initial description from DDL comment
                    "ordinal": ordinal
                })

        # Extract Views
        for m in VIEW_RE.finditer(text):
            schema, name = m.group(1), m.group(2)
            obj_key = f"{schema}.{name}"
            objects[obj_key] = {
                "schema": schema,
                "name": name,
                "type": "VIEW",
                "description": "",
                "source": str(sql_file.name)
            }

        # Extract Comments (COMMENT ON TABLE/COLUMN statements take precedence)
        for m in COMMENT_ON_TABLE_RE.finditer(text):
            schema, name, comment = m.group(1), m.group(2), m.group(3)
            key = f"{schema}.{name}"
            if key in objects: objects[key]["description"] = comment

        for m in COMMENT_ON_COLUMN_RE.finditer(text):
            schema, name, col, comment = m.group(1), m.group(2), m.group(3), m.group(4)
            column_comments[(schema, name, col)] = comment

    # Apply column comments from COMMENT ON COLUMN statements
    for c in columns:
        key = (c["schema"], c["table"], c["column"])
        if key in column_comments:
            c["description"] = column_comments[key]

    return objects, columns


def parse_dbt_sources(objects, columns):
    # Search for dbt schema.yml files
    for yml_file in DBT_ROOT.rglob("*.yml"):
        with open(yml_file, 'r') as f:
            try:
                data = yaml.safe_load(f)
            except yaml.YAMLError:
                continue
            if not data or 'models' not in data: continue
            
            for model in data['models']:
                name = model['name']
                # Determine schema from folder path (e.g., dbt/gold/ -> gold)
                rel_parts = yml_file.relative_to(DBT_ROOT).parts
                schema = rel_parts[0] if rel_parts else "unknown"
                
                obj_key = f"{schema}.{name}"
                if obj_key not in objects:
                    objects[obj_key] = {
                        "schema": schema,
                        "name": name,
                        "type": "DBT_MODEL",
                        "description": model.get('description', ''),
                        "source": str(yml_file.relative_to(ROOT))
                    }
                else:
                    # If SQL object exists, prefer its description unless dbt has one and SQL doesn't
                    if not objects[obj_key]["description"] and model.get('description'):
                        objects[obj_key]["description"] = model.get('description', '')

                if 'columns' in model:
                    for col in model['columns']:
                        col_name = col['name']
                        desc = col.get('description', '')
                        
                        # Update existing column or add new one
                        found = False
                        for c in columns:
                            if c["schema"] == schema and c["table"] == name and c["column"] == col_name:
                                if not c["description"] and desc: c["description"] = desc # Only update if current is empty
                                found = True
                                break
                        if not found:
                            columns.append({
                                "schema": schema,
                                "table": name,
                                "column": col_name,
                                "type": "unknown (dbt)", # dbt yml doesn't specify type directly
                                "nullable": "unknown",
                                "description": desc,
                                "ordinal": 999 # dbt yml doesn't specify ordinal
                            })


def generate_excel(objects, columns):
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "Metadata"
    ws_meta.append(["key", "value"])
    ws_meta.append(["artifact", "FIP_Data_Dictionary.xlsx"])
    ws_meta.append(["generation_time_utc", datetime.now(timezone.utc).isoformat()])
    ws_meta.append(["generation_mode", "generated_from_sql_and_dbt_sources"])
    ws_meta.append(["manual_edit_policy", "DO_NOT_EDIT_MANUALLY_REGENERATE"])

    ws_obj = wb.create_sheet("Objects")
    ws_obj.append(["source", "schema", "object_name", "object_type", "source_path"])
    for r in sorted(objects, key=lambda x: (x["source"], x["schema"], x["object_type"], x["object_name"])):
        ws_obj.append([r["source"], r["schema"], r["object_name"], r["object_type"], r["source_path"]])

    ws_col = wb.create_sheet("Columns")
    ws_col.append(["schema", "table_name", "column_name", "ordinal_position", "data_type", "nullable"])
    for r in sorted(columns, key=lambda x: (x["schema"], x["table"], x["ordinal"])):
        ws_col.append([r["schema"], r["table"], r["column_name"], r["ordinal"], r["type"], r["nullable"]])

    index_rows = locals().get("indexes", globals().get("indexes", []))
    ws_idx = wb.create_sheet("Indexes")
    ws_idx.append(["schema", "table_name", "index_name", "index_columns", "source_path"])
    for r in sorted(index_rows, key=lambda x: (x["schema"], x["table_name"], x["index_name"])):
        ws_idx.append([r["schema"], r["table_name"], r["index_name"], r["index_columns"], r["source_path"]])

    ws_kpi = wb.create_sheet("KPI_Views")
    ws_kpi.append(["schema", "view_name", "source_path"])
    for r in sorted([o for o in objects if o["source"] == "sql" and o["schema"] == "gold" and o["object_type"] == "VIEW" and o["object_name"].startswith("kpi_")], key=lambda x: x["object_name"]):
        ws_kpi.append([r["schema"], r["object_name"], r["source_path"]])

    ws_cfg = wb.create_sheet("Config_Master_Tables")
    ws_cfg.append(["schema", "table_name", "source_path"])
    for r in sorted([o for o in objects if o["source"] == "sql" and o["schema"] == "config" and o["object_type"] == "TABLE"], key=lambda x: x["object_name"]):
        ws_cfg.append([r["schema"], r["object_name"], r["source_path"]])
        
    wb.save(OUTPUT_XLSX)


def generate_markdown(objects, columns):
    with open(OUTPUT_MD, 'w') as f:
        f.write("# FIP Authoritative Data Dictionary\n\n")
        f.write(f"*Generated on: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC*\n\n")
        
        f.write("## Objects Summary\n\n")
        f.write("| Schema | Object Name | Type | Description |\n")
        f.write("| :--- | :--- | :--- | :--- |\n")
        for key in sorted(objects.keys()):
            o = objects[key]
            # Sanitize description for Markdown table
            desc = o["description"].replace("\n", " ").replace("|", "&#124;")
            f.write(f"| {o['schema']} | {o['name']} | {o['type']} | {desc} |\n")
        
        f.write("\n## Column Details\n\n")
        current_obj = ""
        for c in sorted(columns, key=lambda x: (x["schema"], x["table"], x["ordinal"])):
            obj_key = f"{c['schema']}.{c['table']}"
            if obj_key != current_obj:
                f.write(f"\n### {obj_key}\n\n")
                f.write("| Column | Type | Nullable | Description |\n")
                f.write("| :--- | :--- | :--- | :--- |\n")
                current_obj = obj_key
            # Sanitize description for Markdown table
            desc = c["description"].replace("\n", " ").replace("|", "&#124;")
            f.write(f"| {c['column']} | {c['type']} | {c['nullable']} | {desc} |\n")


def main():
    print("Parsing SQL sources...")
    objects, columns = parse_sql_sources()
    print("Parsing dbt sources...")
    parse_dbt_sources(objects, columns)
    
    print(f"Found {len(objects)} objects and {len(columns)} columns.")
    
    print(f"Generating Excel: {OUTPUT_XLSX}")
    generate_excel(objects, columns)
    
    print(f"Generating Markdown: {OUTPUT_MD}")
    generate_markdown(objects, columns)
    
    print("Done.")


if __name__ == "__main__":
    main()
